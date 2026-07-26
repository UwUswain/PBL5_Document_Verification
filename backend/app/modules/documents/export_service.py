import io
import os
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from app.modules.documents.models import Document
from app.modules.users.models import User
from datetime import datetime
from typing import List, Tuple

class ExportService:
    @staticmethod
    async def get_dashboard_stats(db: AsyncSession) -> dict:
        doc_count = await db.execute(select(func.count(Document.id)).where(Document.deleted_at.is_(None)))
        total_documents = doc_count.scalar() or 0

        error_count = await db.execute(
            select(func.count(Document.id)).where(
                (Document.verification_status == 'FAILED') | (Document.status == 'FAILED'),
                Document.deleted_at.is_(None)
            )
        )
        error_documents = error_count.scalar() or 0

        verified_count = await db.execute(
            select(func.count(Document.id)).where(
                Document.verification_status == 'VERIFIED',
                Document.deleted_at.is_(None)
            )
        )
        verified_documents = verified_count.scalar() or 0
        
        suspicious_count = await db.execute(
            select(func.count(Document.id)).where(
                Document.verification_status == 'SUSPICIOUS',
                Document.deleted_at.is_(None)
            )
        )
        suspicious_documents = suspicious_count.scalar() or 0

        return {
            "total": total_documents,
            "error": error_documents,
            "verified": verified_documents,
            "suspicious": suspicious_documents,
            "success_rate": round(((total_documents - error_documents) / total_documents * 100) if total_documents > 0 else 0, 1)
        }

    @staticmethod
    async def get_all_documents(db: AsyncSession) -> List[Tuple[Document, str]]:
        query = select(Document, User.full_name).join(
            User, Document.owner_id == User.id
        ).where(Document.deleted_at.is_(None)).order_by(Document.created_at.desc())
        
        result = await db.execute(query)
        return result.all()

    @staticmethod
    async def export_pdf(db: AsyncSession) -> bytes:
        stats = await ExportService.get_dashboard_stats(db)
        docs = await ExportService.get_all_documents(db)

        # Count categories
        categories = {}
        for doc, _ in docs:
            cat = doc.category or "Khác"
            categories[cat] = categories.get(cat, 0) + 1

        pdf = FPDF()
        pdf.add_page()
        
        # Load fonts
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
        font_regular = os.path.join(base_dir, 'shared', 'fonts', 'Roboto-Regular.ttf')
        font_bold = os.path.join(base_dir, 'shared', 'fonts', 'Roboto-Bold.ttf')
        
        pdf.add_font("Roboto", style="", fname=font_regular, uni=True)
        pdf.add_font("Roboto", style="B", fname=font_bold, uni=True)
        
        # Header
        pdf.set_font("Roboto", style="B", size=24)
        pdf.set_text_color(15, 23, 42)
        pdf.cell(0, 15, "DOCUMIND - BÁO CÁO HỆ THỐNG", ln=True, align="C")
        pdf.ln(5)
        
        pdf.set_font("Roboto", style="", size=12)
        pdf.set_text_color(100, 116, 139)
        pdf.cell(0, 10, f"Thời gian trích xuất: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", ln=True, align="C")
        pdf.ln(10)

        # 1. Thống kê tổng quan
        pdf.set_font("Roboto", style="B", size=16)
        pdf.set_text_color(15, 23, 42)
        pdf.cell(0, 10, "1. Tổng quan hệ thống", ln=True)
        pdf.ln(5)
        
        pdf.set_font("Roboto", style="", size=12)
        pdf.cell(100, 10, f"- Tổng số tài liệu: {stats['total']}", ln=True)
        pdf.cell(100, 10, f"- Tỷ lệ xử lý thành công: {stats['success_rate']}%", ln=True)
        pdf.cell(100, 10, f"- Tài liệu hợp lệ (VERIFIED): {stats['verified']}", ln=True)
        pdf.cell(100, 10, f"- Tài liệu nghi vấn (SUSPICIOUS): {stats['suspicious']}", ln=True)
        pdf.cell(100, 10, f"- Tài liệu lỗi (FAILED): {stats['error']}", ln=True)
        pdf.ln(10)

        # 2. Danh mục tài liệu
        pdf.set_font("Roboto", style="B", size=16)
        pdf.cell(0, 10, "2. Phân loại tài liệu", ln=True)
        pdf.ln(5)
        
        pdf.set_font("Roboto", style="", size=12)
        for cat, count in categories.items():
            pdf.cell(0, 8, f"- {cat.capitalize()}: {count} tài liệu", ln=True)
        pdf.ln(10)

        # 3. Danh sách tài liệu (Preview 20 cái)
        pdf.set_font("Roboto", style="B", size=16)
        pdf.cell(0, 10, "3. Danh sách tài liệu gần đây (Tối đa 20)", ln=True)
        pdf.ln(5)
        
        # Table Header
        pdf.set_font("Roboto", style="B", size=11)
        pdf.set_fill_color(241, 245, 249)
        pdf.cell(90, 10, "Tên tài liệu", border=1, fill=True)
        pdf.cell(40, 10, "Phân loại", border=1, fill=True)
        pdf.cell(30, 10, "Trạng thái", border=1, fill=True)
        pdf.cell(30, 10, "Người đăng", border=1, fill=True)
        pdf.ln()

        # Table rows
        pdf.set_font("Roboto", style="", size=10)
        for doc, full_name in docs[:20]:
            name = doc.file_name[:40] + "..." if len(doc.file_name) > 40 else doc.file_name
            cat = (doc.category or "Khác").capitalize()[:15]
            status = doc.verification_status or doc.status
            owner = full_name.split()[-1] if full_name else "N/A"
            
            pdf.cell(90, 10, name, border=1)
            pdf.cell(40, 10, cat, border=1)
            pdf.cell(30, 10, status, border=1)
            pdf.cell(30, 10, owner, border=1)
            pdf.ln()

        # output as bytearray
        return bytes(pdf.output())

    @staticmethod
    async def export_excel(db: AsyncSession) -> bytes:
        docs = await ExportService.get_all_documents(db)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Document List"
        
        # Headers
        headers = ["STT", "Tên tài liệu", "Phân loại", "Người chia sẻ", "Ngày tạo", "Trạng thái", "Thẩm định AI", "Quyền truy cập"]
        
        # Header styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        # Data
        for row_num, (doc, full_name) in enumerate(docs, 2):
            ai_results = doc.ai_results or {}
            privacy = ai_results.get("privacy_level", "PRIVATE")
            
            ws.cell(row=row_num, column=1, value=row_num - 1).alignment = Alignment(horizontal="center")
            ws.cell(row=row_num, column=2, value=doc.file_name)
            ws.cell(row=row_num, column=3, value=(doc.category or "Khác").capitalize())
            ws.cell(row=row_num, column=4, value=full_name or "N/A")
            ws.cell(row=row_num, column=5, value=doc.created_at.strftime('%d/%m/%Y %H:%M:%S') if doc.created_at else "")
            ws.cell(row=row_num, column=6, value=doc.status)
            ws.cell(row=row_num, column=7, value=doc.verification_status or "N/A")
            ws.cell(row=row_num, column=8, value=privacy)
            
        # Adjust column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 15
        
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
